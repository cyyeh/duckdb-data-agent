import asyncio
import csv
import duckdb
import os
import re
import tempfile
import threading
from typing import Any

SUPPORTED_EXTENSIONS = {".csv", ".json", ".parquet", ".xlsx"}


def _escape_identifier(name: str) -> str:
    """Escape a SQL identifier by doubling any embedded double quotes."""
    return '"' + name.replace('"', '""') + '"'


class Database:
    def __init__(self, db_path: str = ":memory:"):
        self.db_path = db_path
        self.conn = duckdb.connect(db_path)
        self._lock = threading.Lock()

    def close(self) -> None:
        """Close the underlying DuckDB connection."""
        self.conn.close()

    def execute_query(self, sql: str) -> dict[str, Any]:
        """Execute a SQL query and return results as dict with columns, rows, rowCount."""
        result = self.conn.execute(sql)
        if result is None:
            return {"columns": [], "rows": [], "rowCount": 0}
        columns = [desc[0] for desc in result.description] if result.description else []
        rows = [dict(zip(columns, row)) for row in result.fetchall()] if columns else []
        return {"columns": columns, "rows": rows, "rowCount": len(rows)}

    def load_csv(self, file_bytes: bytes, filename: str, table_name: str) -> dict[str, Any]:
        """Load a CSV file into a DuckDB table. Returns table info."""
        ident = _escape_identifier(table_name)
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        try:
            self.conn.execute(
                f"CREATE OR REPLACE TABLE {ident} AS SELECT * FROM read_csv_auto($1)",
                [tmp_path],
            )
        finally:
            os.unlink(tmp_path)
        return self.get_table_info(table_name)

    def load_json(self, file_bytes: bytes, filename: str, table_name: str) -> dict[str, Any]:
        """Load a JSON file (array of objects) into a DuckDB table. Returns table info."""
        ident = _escape_identifier(table_name)
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        try:
            self.conn.execute(
                f"CREATE OR REPLACE TABLE {ident} AS SELECT * FROM read_json_auto($1)",
                [tmp_path],
            )
        finally:
            os.unlink(tmp_path)
        return self.get_table_info(table_name)

    def load_parquet(self, file_bytes: bytes, filename: str, table_name: str) -> dict[str, Any]:
        """Load a Parquet file into a DuckDB table. Returns table info."""
        ident = _escape_identifier(table_name)
        with tempfile.NamedTemporaryFile(suffix=".parquet", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        try:
            self.conn.execute(
                f"CREATE OR REPLACE TABLE {ident} AS SELECT * FROM read_parquet($1)",
                [tmp_path],
            )
        finally:
            os.unlink(tmp_path)
        return self.get_table_info(table_name)

    def load_excel(self, file_bytes: bytes, filename: str, base_table_name: str) -> list[dict[str, Any]]:
        """Load an Excel file into DuckDB tables (one per sheet). Returns list of table infos."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            results = []
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                # Sanitize sheet name for table name
                sanitized_sheet = re.sub(r"[^a-z0-9_]", "_", sheet_name.lower())
                sanitized_sheet = re.sub(r"_+", "_", sanitized_sheet).strip("_")
                table_name = f"{base_table_name}_{sanitized_sheet}" if len(sheet_names) > 1 else base_table_name
                # Write sheet data to temp CSV
                with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="") as csv_tmp:
                    writer = csv.writer(csv_tmp)
                    for row in rows:
                        writer.writerow(row)
                    csv_tmp_path = csv_tmp.name
                try:
                    self.conn.execute(
                        f"CREATE OR REPLACE TABLE {_escape_identifier(table_name)} AS SELECT * FROM read_csv_auto($1)",
                        [csv_tmp_path],
                    )
                    results.append(self.get_table_info(table_name))
                finally:
                    os.unlink(csv_tmp_path)
            wb.close()
            if not results:
                raise ValueError("Excel file contains no data")
            return results
        finally:
            os.unlink(tmp_path)

    def load_file(self, file_bytes: bytes, filename: str, table_name: str) -> list[dict[str, Any]]:
        """Dispatch file loading based on extension. Returns list of table infos."""
        ext = os.path.splitext(filename)[1].lower()
        if ext == ".csv":
            return [self.load_csv(file_bytes, filename, table_name)]
        elif ext == ".json":
            return [self.load_json(file_bytes, filename, table_name)]
        elif ext == ".parquet":
            return [self.load_parquet(file_bytes, filename, table_name)]
        elif ext == ".xlsx":
            return self.load_excel(file_bytes, filename, table_name)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

    def get_table_info(self, table_name: str) -> dict[str, Any]:
        """Get info about a specific table."""
        ident = _escape_identifier(table_name)
        cols_result = self.conn.execute(f"DESCRIBE {ident}")
        columns = [
            {"name": row[0], "type": row[1]}
            for row in cols_result.fetchall()
        ]
        count_result = self.conn.execute(f"SELECT COUNT(*) FROM {ident}")
        row_count = count_result.fetchone()[0]
        return {"name": table_name, "columns": columns, "rowCount": row_count}

    def list_tables(self) -> list[dict[str, Any]]:
        """List all tables with their schema info."""
        tables_result = self.conn.execute("SHOW TABLES")
        table_names = [row[0] for row in tables_result.fetchall()]
        return [self.get_table_info(name) for name in table_names]

    def drop_table(self, table_name: str) -> None:
        """Drop a table."""
        self.conn.execute(f"DROP TABLE IF EXISTS {_escape_identifier(table_name)}")

    def load_sample_data(self, csv_path: str, table_name: str) -> dict[str, Any]:
        """Load sample CSV from a file path."""
        ident = _escape_identifier(table_name)
        self.conn.execute(
            f"CREATE OR REPLACE TABLE {ident} AS SELECT * FROM read_csv_auto($1)",
            [csv_path],
        )
        return self.get_table_info(table_name)

    # -- Async wrappers that offload sync DuckDB work to a thread pool --

    async def _run_sync(self, func, *args, **kwargs):
        """Run a sync method in the default executor to avoid blocking the event loop."""
        def _locked_call():
            with self._lock:
                return func(*args, **kwargs)
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _locked_call)

    async def execute_query_async(self, sql: str) -> dict[str, Any]:
        return await self._run_sync(self.execute_query, sql)

    async def list_tables_async(self) -> list[dict[str, Any]]:
        return await self._run_sync(self.list_tables)

    async def load_file_async(self, file_bytes: bytes, filename: str, table_name: str) -> list[dict[str, Any]]:
        return await self._run_sync(self.load_file, file_bytes, filename, table_name)

    async def load_sample_data_async(self, csv_path: str, table_name: str) -> dict[str, Any]:
        return await self._run_sync(self.load_sample_data, csv_path, table_name)

    async def drop_table_async(self, table_name: str) -> None:
        return await self._run_sync(self.drop_table, table_name)
